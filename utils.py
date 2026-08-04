# -*- coding: utf-8 -*-
"""
프로젝트 공통 설정 및 유틸리티 모음.

- app.py(방문객 등록 페이지)와 pages/1_관리자_페이지.py(관리자 페이지)가 함께 사용하는
  경로 상수, 환경변수 로딩, 방문 기록(visitor_log.csv) 스키마/로딩/수정, 이메일 보고
  테스트 기능, 엑셀 생성 함수를 모아둔다.
- 여기 있는 함수들은 화면(Streamlit UI)에 직접 의존하지 않는 공통 로직만 담당한다.
  방문객 등록 화면 전용 로직(입력값 검증, 서명 저장 등)은 그대로 app.py 에 남겨둔다.
"""

import io
import logging
import os
import re
from datetime import datetime, timedelta, timezone

import pandas as pd
import streamlit as st
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
except ImportError:  # python-dotenv 가 아직 설치되지 않은 환경에서도 앱이 죽지 않게 한다.
    load_dotenv = None

# 실행 위치(현재 작업 디렉터리)와 무관하게 항상 이 파일이 있는 폴더를 기준으로 경로를 잡는다.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

if load_dotenv is not None:
    # 로컬에 .env 파일이 있으면 그 값을 환경변수로 불러온다. (.env 는 .gitignore 로 커밋 제외됨)
    load_dotenv(os.path.join(BASE_DIR, ".env"))

EMPLOYEE_CSV_PATH = os.path.join(BASE_DIR, "employees_dummy.csv")
VISITOR_LOG_PATH = os.path.join(BASE_DIR, "visitor_log.csv")
SIGNATURE_DIR = os.path.join(BASE_DIR, "signatures")
REPORT_RECIPIENTS_CSV_PATH = os.path.join(BASE_DIR, "report_recipients_dummy.csv")

KST = timezone(timedelta(hours=9))  # 한국 표준시(고정 오프셋, 서머타임 없음)

# app.py 와 관리자 페이지가 이 로거를 함께 사용한다. Streamlit 멀티페이지 구조에서는
# 어떤 페이지로 먼저 접속하든 로깅이 항상 설정되도록 여기서 한 번만 구성한다.
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("visitor_system")

# visitor_log.csv 의 정식 열 순서. 새 열이 추가되면 여기에만 추가하면 app.py/관리자 페이지에 반영된다.
VISITOR_LOG_COLUMNS = [
    "visit_id",
    "registered_at",
    "checkout_at",
    "report_sent_at",
    "visitor_name",
    "visitor_company",
    "visitor_phone",
    "vehicle_number",
    "visit_purpose",
    "host_employee_id",
    "host_name",
    "host_department",
    "host_email",
    "visit_location",
    "privacy_consent",
    "signature_path",
]

REPORT_RECIPIENT_COLUMNS = ["recipient_id", "name", "department", "email"]

REPORT_EMAIL_SUBJECT_TEMPLATE = "[테스트] 미발송 방문 차량 기록 {count}건"

# 문자열 끝에 붙은 타임존 표기(Z, +09:00, +0900 등)를 인식하기 위한 패턴.
_TZ_SUFFIX_PATTERN = re.compile(r"(?:Z|[+-]\d{2}:?\d{2})$")


def get_setting(key: str, default: str = "") -> str:
    """환경변수 또는 st.secrets 에서 설정값을 읽어온다. (비밀번호/계정정보를 코드에 두지 않기 위함)"""
    if key in os.environ:
        return os.environ[key]
    try:
        return str(st.secrets.get(key, default))
    except Exception:
        # secrets.toml 이 없는 로컬 테스트 환경에서도 에러 없이 기본값을 사용한다.
        return default


def _strip_timezone_suffix(value):
    """문자열 끝의 타임존 표기를 제거한다. 이 프로젝트의 모든 시각은 KST 벽시계 값으로
    저장되므로, 우연히 타임존 표기가 섞여 있어도 이를 변환하지 않고 그대로 KST 값으로 취급한다."""
    if not isinstance(value, str):
        return value
    return _TZ_SUFFIX_PATTERN.sub("", value.strip())


def _parse_datetime_series(series: pd.Series) -> pd.Series:
    """
    시간대가 없는 문자열과 있는 문자열이 섞여 있어도, 형식이 제각각이어도(초 유무 등)
    오류 없이 datetime 으로 변환한다. 변환할 수 없는 값은 NaT 로 처리된다.
    """
    cleaned = series.astype(str).map(_strip_timezone_suffix)
    return pd.to_datetime(cleaned, errors="coerce", format="mixed")


def load_visitor_log_df() -> pd.DataFrame:
    """
    visitor_log.csv 를 읽어 방문 기록 DataFrame 을 반환한다.

    - 파일이 없거나 비어 있으면 빈 DataFrame(정식 스키마만 있는)을 반환한다.
    - 예전 버전에서 저장되어 일부 열(예: checkout_at, report_sent_at)이 없는 행도
      빈 문자열로 채워 맞춘다.
    - registered_at/checkout_at 을 datetime 으로 변환한 registered_at_dt/checkout_at_dt
      열을 추가로 만든다. 형식이 잘못된 값은 NaT 로 처리되어 화면은 중단되지 않는다.
    - 파일이 다른 프로그램에서 사용 중이라 읽기에 실패하면 RuntimeError 로 원인을 알려준다.
      (호출하는 쪽에서 사용자에게 안내 메시지를 보여주고 재시도할 수 있도록 한다)
    """
    if not os.path.exists(VISITOR_LOG_PATH) or os.path.getsize(VISITOR_LOG_PATH) == 0:
        df = pd.DataFrame(columns=VISITOR_LOG_COLUMNS)
    else:
        try:
            df = pd.read_csv(VISITOR_LOG_PATH, dtype=str, encoding="utf-8-sig")
        except Exception as e:
            raise RuntimeError(
                f"방문 기록 파일(visitor_log.csv)을 읽는 중 오류가 발생했습니다. "
                f"다른 프로그램(엑셀 등)에서 파일을 열어 사용 중인지 확인해주세요. (상세: {e})"
            ) from e
        df.columns = [c.strip() for c in df.columns]

    for column in VISITOR_LOG_COLUMNS:
        if column not in df.columns:
            df[column] = ""
    df = df[VISITOR_LOG_COLUMNS].copy()  # 열 순서 정리 + 스키마에 없는 여분 열 제외

    df = df.fillna("")
    for column in VISITOR_LOG_COLUMNS:
        df[column] = df[column].astype(str).str.strip()

    df["registered_at_dt"] = _parse_datetime_series(df["registered_at"])
    df["checkout_at_dt"] = _parse_datetime_series(df["checkout_at"])

    return df


def _read_current_visitor_log_for_update() -> pd.DataFrame:
    """visitor_log.csv 를 '지금 이 순간' 기준으로 다시 읽어, 수정 전 원본 상태를 반환한다.
    (표에 보이는 필터링된 결과가 아니라 항상 파일 전체를 기준으로 비교/수정하기 위함)"""
    if not os.path.exists(VISITOR_LOG_PATH) or os.path.getsize(VISITOR_LOG_PATH) == 0:
        return pd.DataFrame(columns=VISITOR_LOG_COLUMNS)

    try:
        df = pd.read_csv(VISITOR_LOG_PATH, dtype=str, encoding="utf-8-sig")
    except Exception as e:
        raise RuntimeError(
            f"방문 기록 파일(visitor_log.csv)을 읽는 중 오류가 발생했습니다. "
            f"다른 프로그램(엑셀 등)에서 파일을 열어 사용 중인지 확인해주세요. (상세: {e})"
        ) from e

    df.columns = [c.strip() for c in df.columns]
    for column in VISITOR_LOG_COLUMNS:
        if column not in df.columns:
            df[column] = ""
    df = df[VISITOR_LOG_COLUMNS].copy()
    df = df.fillna("")
    for column in VISITOR_LOG_COLUMNS:
        df[column] = df[column].astype(str).str.strip()
    return df


def save_checkout_times(changes: dict) -> dict:
    """
    관리자 페이지에서 수정한 퇴실일시를 visitor_log.csv 원본에 반영한다.

    - changes: {visit_id: datetime|None|NaT, ...} 형태. 값은 st.data_editor 의
      DatetimeColumn 에서 그대로 전달된 값이라고 가정한다.
    - 항상 파일을 새로 읽어 "현재 저장된 값"과 비교하고, 실제로 달라진 행만 갱신한다.
      (관리자 화면이 검색/필터로 일부만 보여주고 있어도 안전하게 visit_id 기준으로만 반영된다)
    - 저장 값은 초 단위를 제거하고 "YYYY-MM-DD HH:MM" 형식으로 저장한다.
    - datetime/None/NaT 가 아닌 값이 섞여 있으면 그 행은 건너뛰고 skipped 목록에 담는다.
    - 변경된 행이 하나도 없으면 파일을 다시 쓰지 않는다.

    반환값: {"updated_count": int, "skipped": [visit_id, ...]}
    """
    result = {"updated_count": 0, "skipped": []}
    if not changes:
        return result

    current_df = _read_current_visitor_log_for_update()
    if current_df.empty:
        result["skipped"] = list(changes.keys())
        return result

    id_to_index = {vid: idx for idx, vid in zip(current_df.index, current_df["visit_id"])}

    for visit_id, new_value in changes.items():
        row_idx = id_to_index.get(visit_id)
        if row_idx is None:
            result["skipped"].append(visit_id)
            continue

        if pd.isna(new_value):
            new_str = ""
        elif isinstance(new_value, datetime):
            new_str = new_value.strftime("%Y-%m-%d %H:%M")
        else:
            result["skipped"].append(visit_id)
            continue

        old_str = current_df.at[row_idx, "checkout_at"]
        if old_str != new_str:
            current_df.at[row_idx, "checkout_at"] = new_str
            result["updated_count"] += 1

    if result["updated_count"] > 0:
        current_df.to_csv(VISITOR_LOG_PATH, index=False, encoding="utf-8-sig")

    return result


def mark_reports_as_sent(visitor_ids: list, sent_at: datetime) -> int:
    """
    [실제 이메일 발송 성공 시에만 호출] 지정한 visit_id 들의 report_sent_at 을
    sent_at(KST) 값으로 갱신한다.

    현재는 테스트 모드(simulate_report_email)만 사용하므로 이 함수는 어디에서도
    호출하지 않는다. 향후 실제 SMTP 발송 함수를 만들어 발송이 전체 수신자에게
    성공적으로 끝난 뒤에만 이 함수를 호출하도록 연결한다.
    """
    if not visitor_ids:
        return 0

    current_df = _read_current_visitor_log_for_update()
    if current_df.empty:
        return 0

    mask = current_df["visit_id"].isin(visitor_ids)
    updated_count = int(mask.sum())
    if updated_count == 0:
        return 0

    current_df.loc[mask, "report_sent_at"] = sent_at.strftime("%Y-%m-%d %H:%M:%S")
    current_df.to_csv(VISITOR_LOG_PATH, index=False, encoding="utf-8-sig")
    return updated_count


def get_unsent_records() -> pd.DataFrame:
    """
    report_sent_at 이 비어 있는(아직 이메일 보고에 포함되지 않은) 방문 기록만 반환한다.
    관리자 화면의 검색어/날짜/회사명 등 현재 필터와는 무관하게, 항상 visitor_log.csv
    전체를 기준으로 조회한다. (늦게 등록된 방문 기록이 누락되지 않도록 하기 위함)
    """
    log_df = load_visitor_log_df()
    return log_df[log_df["report_sent_at"] == ""].copy()


def load_report_recipients() -> pd.DataFrame:
    """
    report_recipients_dummy.csv(이메일 보고 수신자 더미 명단)를 읽어 반환한다.
    파일이 없거나 필수 열이 없거나 읽는 중 오류가 나면, 화면에서 안내할 수 있도록
    RuntimeError 로 원인을 알려준다. (앱이 중단되지 않고 호출부에서 st.error 로 표시)
    """
    if not os.path.exists(REPORT_RECIPIENTS_CSV_PATH):
        raise RuntimeError(
            f"이메일 수신자 명단 파일({os.path.basename(REPORT_RECIPIENTS_CSV_PATH)})을 "
            "찾을 수 없습니다. 프로젝트 폴더에 파일이 있는지 확인해주세요."
        )

    try:
        df = pd.read_csv(REPORT_RECIPIENTS_CSV_PATH, dtype=str, encoding="utf-8-sig")
    except Exception as e:
        raise RuntimeError(f"이메일 수신자 명단 파일을 읽는 중 오류가 발생했습니다: {e}") from e

    df.columns = [c.strip() for c in df.columns]
    missing_columns = [c for c in REPORT_RECIPIENT_COLUMNS if c not in df.columns]
    if missing_columns:
        raise RuntimeError(
            "이메일 수신자 명단 파일에 다음 필수 열이 없습니다: "
            + ", ".join(missing_columns)
            + f" (필요한 열: {', '.join(REPORT_RECIPIENT_COLUMNS)})"
        )

    for column in REPORT_RECIPIENT_COLUMNS:
        df[column] = df[column].fillna("").astype(str).str.strip()

    return df[REPORT_RECIPIENT_COLUMNS]


def build_report_attachment_filename(executed_at: datetime) -> str:
    """이메일 첨부 엑셀 파일명을 만든다. 예: 방문차량기록_2026-08-04.xlsx"""
    return f"방문차량기록_{executed_at.strftime('%Y-%m-%d')}.xlsx"


def simulate_report_email(
    recipients: list,
    unsent_count: int,
    vehicle_count: int,
    executed_at: datetime,
) -> list:
    """
    [테스트 모드 전용] 실제 SMTP 연결/발송을 전혀 하지 않고, 수신자별로 발송될 내용을
    만들어 반환하고 서버 로그에도 남긴다. report_sent_at 은 절대 변경하지 않는다.

    recipients: {"name":..., "department":..., "email":...} 형태의 dict 리스트
    반환값: 화면 표시에 바로 쓸 수 있는 dict 리스트
    """
    subject = REPORT_EMAIL_SUBJECT_TEMPLATE.format(count=unsent_count)
    attachment_filename = build_report_attachment_filename(executed_at)
    executed_at_str = executed_at.strftime("%Y-%m-%d %H:%M:%S")

    results = []
    for recipient in recipients:
        info = {
            "recipient_name": recipient["name"],
            "recipient_department": recipient["department"],
            "recipient_email": recipient["email"],
            "subject": subject,
            "attachment_filename": attachment_filename,
            "executed_at": executed_at_str,
            "unsent_count": unsent_count,
            "vehicle_count": vehicle_count,
        }
        logger.info(
            "[TEST REPORT EMAIL] 수신자=%s(%s) <%s> 제목=%s 첨부파일=%s "
            "실행시각=%s 대상건수=%d 차량수=%d (테스트 모드: 실제 발송 안 함)",
            recipient["name"],
            recipient["department"],
            recipient["email"],
            subject,
            attachment_filename,
            executed_at_str,
            unsent_count,
            vehicle_count,
        )
        results.append(info)

    return results


def build_excel_bytes(
    source_df: pd.DataFrame,
    columns: list,
    column_labels: dict,
    sheet_name: str = "방문기록",
    datetime_column_key: str = "registered_at",
    datetime_source_col: str = "registered_at_dt",
    totals_row: dict = None,
) -> bytes:
    """
    DataFrame 을 서식(굵은 헤더, 자동 필터, 첫 행 고정, 열 너비 자동조정, 날짜 서식)이
    적용된 .xlsx 파일 바이트로 변환한다.

    - columns: 내보낼 열(원본 컬럼명) 순서
    - column_labels: 원본 컬럼명 -> 화면/엑셀에 표시할 한글 라벨
    - datetime_column_key 는 columns 안에 포함된, 날짜/시간으로 표시할 열의 원본 컬럼명이며
      datetime_source_col(예: registered_at_dt) 의 실제 datetime 값을 사용한다.
      변환 실패(NaT)한 값은 정보 손실을 막기 위해 원본 문자열 그대로 남겨둔다.
    - totals_row: {한글라벨: 표시값} 형태로 지정하면, 데이터 마지막 줄 아래에 합계 행을
      굵게+상단 테두리+배경색을 적용해 추가한다. 지정하지 않으면 합계 행 없이 기존과 동일하게 동작한다.
      합계 행은 자동 필터 범위와 날짜 서식 적용 대상에서 제외된다.
    """
    export_df = source_df[columns].copy()

    if datetime_column_key in export_df.columns and datetime_source_col in source_df.columns:
        def resolve_datetime(row):
            dt_value = row[datetime_source_col]
            return dt_value.to_pydatetime() if pd.notna(dt_value) else row[datetime_column_key]

        export_df[datetime_column_key] = source_df.apply(resolve_datetime, axis=1)

    export_df = export_df.rename(columns=column_labels)

    data_row_count = len(export_df)
    totals_values = None
    if totals_row:
        totals_values = [totals_row.get(col, "") for col in export_df.columns]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        # 첫 번째 행 굵게 표시 + 고정
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        worksheet.freeze_panes = "A2"

        # 자동 필터는 실제 데이터 범위까지만 적용한다(합계 행은 필터 대상에서 제외).
        last_col_letter = get_column_letter(len(export_df.columns))
        worksheet.auto_filter.ref = f"A1:{last_col_letter}{data_row_count + 1}"

        # 날짜/시간 열 서식 통일 (데이터 행에만 적용, 합계 행은 제외)
        datetime_label = column_labels.get(datetime_column_key)
        if datetime_label in list(export_df.columns):
            col_idx = list(export_df.columns).index(datetime_label) + 1
            col_letter = get_column_letter(col_idx)
            for row_idx in range(2, data_row_count + 2):
                worksheet[f"{col_letter}{row_idx}"].number_format = "yyyy-mm-dd hh:mm:ss"

        # 합계 행 추가 (요청 시): 굵게 + 상단 테두리 + 배경색으로 일반 데이터 행과 구분한다.
        if totals_values is not None:
            worksheet.append(totals_values)
            totals_row_idx = worksheet.max_row
            top_border = Border(top=Side(style="thin"))
            totals_fill = PatternFill(fill_type="solid", start_color="FFE7E6E6", end_color="FFE7E6E6")
            for cell in worksheet[totals_row_idx]:
                cell.font = Font(bold=True)
                cell.border = top_border
                cell.fill = totals_fill

        # 열 너비 자동 조정 (합계 행 내용의 길이도 함께 고려)
        for idx, column_name in enumerate(export_df.columns, start=1):
            col_letter2 = get_column_letter(idx)
            candidates = [len(str(column_name))] + [len(str(v)) for v in export_df[column_name].tolist()]
            if totals_values is not None:
                candidates.append(len(str(totals_values[idx - 1])))
            max_len = max(candidates)
            worksheet.column_dimensions[col_letter2].width = min(max_len + 4, 40)

    return buffer.getvalue()
